import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from django.http import Http404, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Applicant, Session, AttendanceRecord


# ─── Helpers ────────────────────────────────────────────────────────────────

def is_admin(user):
    return user.is_staff


def _close_expired_sessions():
    """
    Close any active sessions whose end_time has passed.
    Called at the top of every admin-facing view so the status
    is always accurate when the admin looks at the dashboard.
    """
    now_time = timezone.localtime(timezone.now()).time()
    Session.objects.filter(
        is_active=True,
        end_time__isnull=False,
        end_time__lte=now_time,
    ).update(is_active=False)


def generate_user_id():
    """
    Generate the next sequential user ID in the format DMG_CIC_001.
    Finds the highest existing number and increments by 1.
    """
    prefix = 'DMG_CIC_'
    existing = (
        User.objects.filter(username__startswith=prefix)
        .values_list('username', flat=True)
    )
    max_num = 0
    for username in existing:
        try:
            num = int(username.replace(prefix, ''))
            if num > max_num:
                max_num = num
        except ValueError:
            pass
    return f"{prefix}{str(max_num + 1).zfill(3)}"


# ─── Public pages ────────────────────────────────────────────────────────────

def home_view(request):
    return render(request, 'core/home.html')


def login_view(request):
    if request.user.is_authenticated:
        return _redirect_after_login(request.user)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return _redirect_after_login(user)
        messages.error(request, 'Invalid username or password.')

    return render(request, 'core/login.html')


def logout_view(request):
    logout(request)
    return redirect('home')


def register_view(request):
    if request.user.is_authenticated:
        return _redirect_after_login(request.user)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        phone      = request.POST.get('phone', '').strip()
        cohort     = request.POST.get('cohort', '').strip()
        password   = request.POST.get('password', '')
        password2  = request.POST.get('password2', '')

        # Basic validation
        if not first_name or not last_name:
            messages.error(request, 'First name and last name are required.')
            return render(request, 'core/register.html', {'post': request.POST, 'cohort_choices': Applicant.COHORT_CHOICES})

        if len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'core/register.html', {'post': request.POST, 'cohort_choices': Applicant.COHORT_CHOICES})

        if password != password2:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'core/register.html', {'post': request.POST, 'cohort_choices': Applicant.COHORT_CHOICES})

        # Generate unique user ID
        user_id = generate_user_id()

        user = User.objects.create_user(
            username=user_id,
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
        Applicant.objects.create(user=user, phone=phone, cohort=cohort)

        # Show the generated ID before redirecting to login
        messages.success(
            request,
            f'Registration successful! Your User ID is: {user_id} — save this, you will need it to sign in.'
        )
        return redirect('login')

    return render(request, 'core/register.html', {'cohort_choices': Applicant.COHORT_CHOICES})


def _redirect_after_login(user):
    if user.is_staff:
        return redirect('admin_dashboard')
    return redirect('applicant_dashboard')


# ─── QR Check-in ─────────────────────────────────────────────────────────────

@login_required(login_url='login')
def qr_checkin_view(request, token):
    """
    QR check-in — requires the applicant to be logged in.
    They can only check in themselves; no username field is accepted.
    """
    session = get_object_or_404(Session, qr_token=token, is_active=True)

    # Admins don't have an applicant profile — redirect them away
    if request.user.is_staff:
        messages.info(request, 'Admins do not check in. Use the admin panel to manage attendance.')
        return redirect('admin_dashboard')

    try:
        applicant = request.user.applicant
    except Applicant.DoesNotExist:
        messages.error(request, 'No applicant profile found for your account.')
        return redirect('home')

    # Real-time end-time guard (catches the gap before the scheduler fires)
    now_time = timezone.localtime(timezone.now()).time()
    if session.end_time and now_time >= session.end_time:
        session.is_active = False
        session.save(update_fields=['is_active'])
        return render(request, 'core/qr_checkin.html', {'session_ended': True})

    already_checked_in = AttendanceRecord.objects.filter(
        applicant=applicant, session=session
    ).exists()

    if request.method == 'POST':
        if already_checked_in:
            messages.warning(request, 'You have already checked in for this session.')
            return render(request, 'core/qr_checkin.html', {
                'session': session, 'already_checked_in': True
            })

        now_time = timezone.localtime(timezone.now()).time()
        status = 'late' if now_time > session.late_after else 'present'
        AttendanceRecord.objects.create(applicant=applicant, session=session, status=status)

        return render(request, 'core/qr_checkin.html', {
            'session': session,
            'checked_in': True,
            'status': status,
            'applicant': applicant,
        })

    return render(request, 'core/qr_checkin.html', {
        'session': session,
        'applicant': applicant,
        'already_checked_in': already_checked_in,
    })


# ─── Manual fallback check-in (no QR) ────────────────────────────────────────

@login_required(login_url='login')
def manual_checkin_view(request):
    """
    Fallback check-in — requires login.
    Applicant checks in only themselves; no username field accepted.
    """
    if request.user.is_staff:
        messages.info(request, 'Use the admin panel to manage attendance manually.')
        return redirect('admin_dashboard')

    try:
        applicant = request.user.applicant
    except Applicant.DoesNotExist:
        messages.error(request, 'No applicant profile found for your account.')
        return redirect('home')

    active_session = Session.objects.filter(is_active=True).first()

    # Real-time end-time guard
    if active_session and active_session.end_time:
        now_time = timezone.localtime(timezone.now()).time()
        if now_time >= active_session.end_time:
            active_session.is_active = False
            active_session.save(update_fields=['is_active'])
            active_session = None

    already_checked_in = (
        active_session and
        AttendanceRecord.objects.filter(applicant=applicant, session=active_session).exists()
    )

    if request.method == 'POST':
        if not active_session:
            messages.error(request, 'No active session right now. Please ask your trainer.')
            return render(request, 'core/manual_checkin.html', {'session': None, 'applicant': applicant})

        if already_checked_in:
            messages.warning(request, 'You have already checked in for this session.')
            return render(request, 'core/manual_checkin.html', {
                'session': active_session, 'applicant': applicant, 'already_checked_in': True
            })

        now_time = timezone.localtime(timezone.now()).time()
        status = 'late' if now_time > active_session.late_after else 'present'
        AttendanceRecord.objects.create(applicant=applicant, session=active_session, status=status)

        return render(request, 'core/manual_checkin.html', {
            'session': active_session,
            'checked_in': True,
            'status': status,
            'applicant': applicant,
        })

    return render(request, 'core/manual_checkin.html', {
        'session': active_session,
        'applicant': applicant,
        'already_checked_in': already_checked_in,
    })


# ─── Applicant portal ────────────────────────────────────────────────────────

@login_required(login_url='login')
def applicant_dashboard_view(request):
    try:
        applicant = request.user.applicant
    except Applicant.DoesNotExist:
        messages.error(request, 'No applicant profile found for your account.')
        return redirect('home')

    _close_expired_sessions()   # keep QR card state accurate
    records = applicant.attendance_records.select_related('session').all()
    summary = applicant.attendance_summary()
    active_session = Session.objects.filter(is_active=True).first()

    return render(request, 'core/applicant_dashboard.html', {
        'applicant': applicant,
        'records': records,
        'summary': summary,
        'active_session': active_session,
    })


# ─── Admin dashboard ─────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def admin_dashboard_view(request):
    _close_expired_sessions()   # ensure stale sessions are closed before rendering
    sessions = Session.objects.all()
    applicants = Applicant.objects.select_related('user').all()
    active_session = Session.objects.filter(is_active=True).first()

    context = {
        'sessions': sessions,
        'applicants': applicants,
        'active_session': active_session,
        'total_applicants': applicants.count(),
    }
    if active_session:
        context['stats'] = active_session.attendance_stats()

    return render(request, 'core/admin_dashboard.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def session_detail_view(request, session_id):
    _close_expired_sessions()   # ensure stale sessions are closed before rendering
    session = get_object_or_404(Session, pk=session_id)
    records = session.attendance_records.select_related('applicant__user').all()

    # Build a list of all applicants with their record (or absent) for this session
    all_applicants = Applicant.objects.select_related('user').all()
    checked_in_ids = records.values_list('applicant_id', flat=True)

    attendance_list = []
    for applicant in all_applicants:
        record = records.filter(applicant=applicant).first()
        attendance_list.append({
            'applicant': applicant,
            'record': record,
            'status': record.status if record else 'absent',
        })

    return render(request, 'core/session_detail.html', {
        'session': session,
        'attendance_list': attendance_list,
        'stats': session.attendance_stats(),
    })


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def mark_absent_view(request, session_id):
    """Admin manually marks all unchecked applicants as absent for a session."""
    session = get_object_or_404(Session, pk=session_id)
    if request.method == 'POST':
        all_applicants = Applicant.objects.all()
        for applicant in all_applicants:
            AttendanceRecord.objects.get_or_create(
                applicant=applicant,
                session=session,
                defaults={'status': 'absent'}
            )
        messages.success(request, 'All unchecked applicants have been marked absent.')
    return redirect('session_detail', session_id=session_id)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def admin_mark_attendance_view(request, session_id, applicant_id):
    """Admin manually sets attendance status for a single applicant."""
    session   = get_object_or_404(Session, pk=session_id)
    applicant = get_object_or_404(Applicant, pk=applicant_id)

    if request.method == 'POST':
        status = request.POST.get('status', '').strip()
        if status not in ('present', 'late', 'absent'):
            messages.error(request, 'Invalid status.')
            return redirect('session_detail', session_id=session_id)

        record, created = AttendanceRecord.objects.get_or_create(
            applicant=applicant,
            session=session,
            defaults={'status': status},
        )
        if not created:
            record.status = status
            record.save()

        action = 'created' if created else 'updated'
        messages.success(
            request,
            f'{applicant} marked as {status.capitalize()} ({action}).'
        )
    return redirect('session_detail', session_id=session_id)


# ─── Export helpers ──────────────────────────────────────────────────────────

def _build_attendance_rows(session):
    """Return header + data rows for a session's attendance list."""
    all_applicants = Applicant.objects.select_related('user').all()
    records = session.attendance_records.select_related('applicant__user').all()

    headers = ['#', 'User ID', 'Full Name', 'Cohort', 'Phone', 'Status', 'Check-in Time']
    rows = []
    for i, applicant in enumerate(all_applicants, start=1):
        record = records.filter(applicant=applicant).first()
        status = record.status.capitalize() if record else 'Absent'
        checkin_time = (
            record.checked_in_at.strftime('%H:%M') if record and record.status != 'absent' else '—'
        )
        rows.append([
            i,
            applicant.user.username,
            applicant.user.get_full_name() or '—',
            applicant.cohort or '—',
            applicant.phone or '—',
            status,
            checkin_time,
        ])
    return headers, rows


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def export_csv_view(request, session_id):
    session = get_object_or_404(Session, pk=session_id)
    headers, rows = _build_attendance_rows(session)

    filename = f"attendance_{session.date}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    # Session info header
    writer.writerow([f'NCA Digital Skills — Attendance Report'])
    writer.writerow([f'Session: {session.date} — {session.title or "No title"}'])
    writer.writerow([f'Exported: {timezone.now().strftime("%Y-%m-%d %H:%M")}'])
    writer.writerow([])
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return response


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
def export_excel_view(request, session_id):
    session = get_object_or_404(Session, pk=session_id)
    headers, rows = _build_attendance_rows(session)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {session.date}"

    # ── Colour palette ──────────────────────────────────────────────────────
    dark    = '3D3520'
    amber   = 'D4821A'
    sand    = 'E8D5A3'
    cream   = 'F5F5E0'
    white   = 'FFFFFF'
    green   = 'D1FAE5'
    green_t = '065F46'
    yellow  = 'FEF3C7'
    yellow_t= '92400E'
    red     = 'FEE2E2'
    red_t   = '991B1B'

    thin_border = Border(
        left=Side(style='thin', color='D4821A'),
        right=Side(style='thin', color='D4821A'),
        top=Side(style='thin', color='D4821A'),
        bottom=Side(style='thin', color='D4821A'),
    )

    # ── Title block ─────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'NCA Digital Skills Centre — Attendance Report'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color=white)
    title_cell.fill = PatternFill('solid', fgColor=dark)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    sub_cell = ws['A2']
    sub_cell.value = f"Session: {session.date}  |  {session.title or 'No title'}  |  Exported: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font = Font(name='Calibri', size=10, color=dark)
    sub_cell.fill = PatternFill('solid', fgColor=sand)
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.append([])  # blank row 3

    # ── Column headers (row 4) ───────────────────────────────────────────────
    ws.append(headers)
    header_row = ws[4]
    for cell in header_row:
        cell.font = Font(name='Calibri', bold=True, size=10, color=white)
        cell.fill = PatternFill('solid', fgColor=amber)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[4].height = 20

    # ── Data rows ────────────────────────────────────────────────────────────
    for idx, row in enumerate(rows):
        ws.append(row)
        excel_row = ws[idx + 5]
        status_val = str(row[5]).lower()

        if status_val == 'present':
            row_fill = PatternFill('solid', fgColor=green)
            status_font_color = green_t
        elif status_val == 'late':
            row_fill = PatternFill('solid', fgColor=yellow)
            status_font_color = yellow_t
        else:
            row_fill = PatternFill('solid', fgColor=red)
            status_font_color = red_t

        for col_idx, cell in enumerate(excel_row, start=1):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=10)
            # Alternate row shading for non-status rows
            if col_idx != 6:
                cell.fill = PatternFill('solid', fgColor=cream if idx % 2 == 0 else white)
            else:
                cell.fill = row_fill
                cell.font = Font(name='Calibri', size=10, bold=True, color=status_font_color)

        ws.row_dimensions[idx + 5].height = 18

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [5, 18, 24, 16, 16, 12, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Freeze panes below header ────────────────────────────────────────────
    ws.freeze_panes = 'A5'

    # ── Write response ───────────────────────────────────────────────────────
    filename = f"attendance_{session.date}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
